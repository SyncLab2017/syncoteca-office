"""Catalog anomaly detection: tracks with link but missing metadata fields."""
import os
import httpx


_REQUIRED_FIELDS = ["artist", "music_author", "lyrics_author", "label"]
_FIELD_LABELS = {
    "artist": "Исполнитель",
    "music_author": "Автор музыки",
    "lyrics_author": "Автор слов",
    "label": "Лейбл",
}


def _sb_headers() -> dict:
    key = os.getenv("SUPABASE_KEY", "")
    return {"apikey": key, "Authorization": f"Bearer {key}"}

def _sb_base() -> str:
    return os.getenv("SUPABASE_URL", "").rstrip("/")


def fetch_anomalies(limit: int = 2000) -> list[dict]:
    """Return tracks where link IS NOT NULL but at least one key field is missing."""
    params = {
        "select": "id,title,artist,music_author,lyrics_author,label,link,release_date",
        "link": "not.is.null",
        "or": f"({','.join(f'{f}.is.null' for f in _REQUIRED_FIELDS)})",
        "order": "id.asc",
        "limit": str(limit),
    }
    r = httpx.get(
        f"{_sb_base()}/rest/v1/tracks",
        headers=_sb_headers(),
        params=params,
        timeout=15,
    )
    r.raise_for_status()
    return r.json()


def analyze_anomalies(tracks: list[dict]) -> dict:
    """Break down anomalies by missing field combination."""
    by_missing: dict[str, list[dict]] = {}

    for t in tracks:
        missing = [f for f in _REQUIRED_FIELDS if not t.get(f)]
        if not missing:
            continue
        key = "+".join(missing)
        by_missing.setdefault(key, []).append(t)

    return by_missing


def format_audit_report(tracks: list[dict], sample_size: int = 5) -> str:
    """Format a Telegram-friendly anomaly report."""
    if not tracks:
        return "✅ Ковальски: аномалий не найдено — у всех треков со ссылками метаданные заполнены."

    breakdown = analyze_anomalies(tracks)
    lines = [f"🔍 Ковальски: найдено {len(tracks)} треков со ссылками, но неполными метаданными\n"]

    for combo_key, items in sorted(breakdown.items(), key=lambda x: -len(x[1])):
        labels = [_FIELD_LABELS[f] for f in combo_key.split("+")]
        lines.append(f"❌ Нет [{', '.join(labels)}]: {len(items)} треков")
        for t in items[:sample_size]:
            title = t.get("title") or "—"
            artist = t.get("artist") or "?"
            link = t.get("link") or ""
            lines.append(f"  • «{title}» — {artist} | {link[:50]}")
        if len(items) > sample_size:
            lines.append(f"  … и ещё {len(items) - sample_size}")
        lines.append("")

    lines.append("💡 Действия:")
    lines.append("  • /enrich — запустить авто-обогащение через скрипт")
    lines.append("  • /export_anomalies — выгрузить список в Excel для ручной правки")

    return "\n".join(lines)


def run_audit(limit: int = 2000) -> tuple[list[dict], str]:
    """Fetch anomalies and return (tracks, formatted_report)."""
    tracks = fetch_anomalies(limit)
    report = format_audit_report(tracks)
    return tracks, report


def export_anomalies_excel(tracks: list[dict]) -> bytes:
    """Build an Excel workbook of anomalous tracks."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Аномалии"

    header_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    columns = [
        ("ID", 8),
        ("Название", 40),
        ("Исполнитель", 25),
        ("Автор музыки", 25),
        ("Автор слов", 25),
        ("Лейбл", 20),
        ("Отсутствует", 35),
        ("Ссылка", 50),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = col_width

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    for i, t in enumerate(tracks, 2):
        missing = [_FIELD_LABELS[f] for f in _REQUIRED_FIELDS if not t.get(f)]
        ws.cell(row=i, column=1, value=t.get("id"))
        ws.cell(row=i, column=2, value=t.get("title") or "")
        ws.cell(row=i, column=3, value=t.get("artist") or "")
        ws.cell(row=i, column=4, value=t.get("music_author") or "")
        ws.cell(row=i, column=5, value=t.get("lyrics_author") or "")
        ws.cell(row=i, column=6, value=t.get("label") or "")
        ws.cell(row=i, column=7, value=", ".join(missing))
        ws.cell(row=i, column=8, value=t.get("link") or "")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
