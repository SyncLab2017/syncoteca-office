"""Excel export of Supabase tracks catalog with natural-language filter parsing."""
import datetime
import io
import os
import re

import httpx
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def _sb_headers() -> dict:
    key = os.getenv("SUPABASE_KEY", "")
    return {"apikey": key, "Authorization": f"Bearer {key}"}

def _sb_base() -> str:
    return os.getenv("SUPABASE_URL", "").rstrip("/")


# Short label aliases: when Denis says just "Мелодия" or "Зион", resolve to label search.
# Keys are lowercase. Values are the search term passed to label.ilike.*value*.
LABEL_ALIASES: dict[str, str] = {
    "мелодия": "Мелодия",
    "melody": "Мелодия",
    "зион": "Zion",
    "zion": "Zion",
    "джем": "ДЖЕМ",
    "auris": "Auris",
    "аурис": "Auris",
    "dnk": "DNK",
    "днк": "DNK",
    "adam": "Adam Music",
    "balt": "Balt",
    "golden sound": "Golden Sound",
}

# Words that signal "label context" — "фирме Мелодия", "компании X", "издательства X"
_LABEL_CONTEXT_RE = re.compile(
    r"(?:лейбл[аеыуой]?|label|фирм[аеыуой]|компани[яи]|издательств[аоеу]?|"
    r"правообладател[яьей]+|рекорд-лейбл[аеыуой]?)\s+([«»\w\s.,-]+?)(?:\s+(?:год|за|и|дай|скинь)|[.,!?]|$)",
    re.IGNORECASE,
)

# Genre keywords: Russian/English bare words that map to genre_1 search terms
GENRE_ALIASES: dict[str, str] = {
    "шансон": "шансон", "chanson": "шансон", "shanson": "шансон",
    "джаз": "джаз", "jazz": "джаз",
    "рок": "рок", "rock": "рок",
    "поп": "поп", "pop": "поп",
    "классика": "классик", "classical": "классик", "классическая": "классик",
    "электронная": "electro", "электро": "electro", "electronic": "electro",
    "хип-хоп": "hip", "хипхоп": "hip", "hip-hop": "hip",
    "рэп": "rap", "rap": "rap",
    "фолк": "folk", "folk": "folk",
    "блюз": "blues", "blues": "blues",
    "кантри": "country", "country": "country",
    "металл": "metal", "metal": "metal",
    "инструментальная": "instrumental", "instrumental": "instrumental",
    "ambient": "ambient", "эмбиент": "ambient",
    "ретро": "ретро", "retro": "ретро",
    "советская": "советск", "советский": "советск",
}

_GENRE_CONTEXT_RE = re.compile(
    r"(?:жанр[аеыуой]?|стил[еяьи]?|категори[яи]|genre|style|category)\s+([\w\s-]+?)(?:[.,!?]|$)",
    re.IGNORECASE,
)


def _strip_quotes(s: str) -> str:
    return re.sub(r'[«»"\']', '', s).strip()


def parse_export_query(text: str) -> dict:
    """Extract filters from natural-language export request.

    Returns dict with optional keys: artist, label, year_from, year_to, genre
    Examples:
      "репертуар S.T.A.L.K.E.R." → {"artist": "S.T.A.L.K.E.R."}
      "треки за 1996 год"         → {"year_from": 1996, "year_to": 1996}
      "период 1975-1980"          → {"year_from": 1975, "year_to": 1980}
      "лейбл Мелодия"             → {"label": "Мелодия"}
      "по фирме «Мелодия»"        → {"label": "Мелодия"}
      "Мелодия"                   → {"label": "Мелодия"}  (via alias)
    """
    filters: dict = {}
    lower = text.lower()

    # === Specific date detection (must run before year parsing) ===
    # Detect "вчера", "сегодня", "DD.MM.YYYY", "DD.MM", Russian months
    # Sets filters["date_added"] (created_at) or filters["release_day"] (release_date)
    _today = datetime.date.today()
    _is_release_ctx = any(w in lower for w in ("вышл", "релиз", "вышедш", "release", "выход"))
    _detected_date: datetime.date | None = None

    if "вчера" in lower or "yesterday" in lower:
        _detected_date = _today - datetime.timedelta(days=1)
    elif "позавчера" in lower:
        _detected_date = _today - datetime.timedelta(days=2)
    elif "сегодня" in lower or "today" in lower:
        _detected_date = _today
    else:
        # DD.MM.YYYY
        _dm = re.search(r'\b(\d{1,2})\.(\d{2})\.(\d{4})\b', text)
        if _dm:
            try:
                _detected_date = datetime.date(int(_dm.group(3)), int(_dm.group(2)), int(_dm.group(1)))
                text = (text[:_dm.start()] + text[_dm.end():]).strip()
                lower = text.lower()
            except ValueError:
                pass
        # YYYY-MM-DD (ISO, for round-trip re-parse from cached filters)
        if not _detected_date:
            _dm = re.search(r'\b(\d{4})-(\d{2})-(\d{2})\b', text)
            if _dm:
                try:
                    _detected_date = datetime.date(int(_dm.group(1)), int(_dm.group(2)), int(_dm.group(3)))
                    text = (text[:_dm.start()] + text[_dm.end():]).strip()
                    lower = text.lower()
                except ValueError:
                    pass
        # DD.MM (current year)
        if not _detected_date:
            _dm = re.search(r'\b(\d{1,2})\.(\d{2})\b(?!\.\d)', text)
            if _dm:
                _dv, _mv = int(_dm.group(1)), int(_dm.group(2))
                if 1 <= _dv <= 31 and 1 <= _mv <= 12:
                    try:
                        _detected_date = datetime.date(_today.year, _mv, _dv)
                        text = (text[:_dm.start()] + text[_dm.end():]).strip()
                        lower = text.lower()
                    except ValueError:
                        pass
        # Russian month names: "10 июня", "10 июня 2026"
        if not _detected_date:
            _RU_MONTHS = [
                ("январ", 1), ("феврал", 2), ("март", 3), ("апрел", 4),
                (r"ма[йя]", 5), ("июн", 6), ("июл", 7), ("август", 8),
                ("сентябр", 9), ("октябр", 10), ("ноябр", 11), ("декабр", 12),
            ]
            for _mp, _mn in _RU_MONTHS:
                _dm = re.search(r'(\d{1,2})\s+' + _mp + r'\w*(?:\s+(\d{4}))?', lower)
                if _dm:
                    try:
                        _yr = int(_dm.group(2)) if _dm.group(2) else _today.year
                        _detected_date = datetime.date(_yr, _mn, int(_dm.group(1)))
                        text = (text[:_dm.start()] + text[_dm.end():]).strip()
                        lower = text.lower()
                        break
                    except ValueError:
                        pass

    if _detected_date:
        _date_key = "release_day" if _is_release_ctx else "date_added"
        filters[_date_key] = _detected_date.isoformat()

    # Handle CLI-style flags: --label="X" / --artist="X" (user copied bot suggestion)
    m = re.search(r'--label=["\']?([^"\']+)["\']?', text, re.IGNORECASE)
    if m:
        filters["label"] = _strip_quotes(m.group(1).strip())
        return filters
    m = re.search(r'--artist=["\']?([^"\']+)["\']?', text, re.IGNORECASE)
    if m:
        filters["artist"] = _strip_quotes(m.group(1).strip())
        return filters

    # Year range: "1975-1980" / "1975–1980"
    # Do NOT return early — continue to extract artist/label from the same query
    # so "Барыкин 1990-2000" gets both artist AND year_from/year_to.
    m = re.search(r"\b((?:19|20)\d{2})\s*[-–—]\s*((?:19|20)\d{2})\b", text)
    if m:
        filters["year_from"] = int(m.group(1))
        filters["year_to"] = int(m.group(2))
        # Strip year range from text so it doesn't pollute artist detection below
        text = (text[:m.start()] + text[m.end():]).strip()
        lower = text.lower()

    # Single year (only if no range already found)
    if not filters.get("year_from"):
        m = re.search(r"\b((?:19|20)\d{2})\b", text)
        if m:
            filters["year_from"] = int(m.group(1))
            filters["year_to"] = int(m.group(1))
            text = (text[:m.start()] + text[m.end():]).strip()
            lower = text.lower()

    # Label: "лейбл X" / "label X" / "фирме X" / "компании X" / "издательства X"
    m = _LABEL_CONTEXT_RE.search(text)
    if m:
        filters["label"] = _strip_quotes(m.group(1).strip())

    # Quoted artist name: «Руки Вверх», "Аквариум" — guillemets/quotes signal artist directly
    if not filters.get("artist"):
        m = re.search(r'[«"]([\w\s.\-!?]+?)[»"]', text)
        if m:
            candidate = m.group(1).strip()
            # Accept if 1–5 words and not a genre/label/year phrase
            cand_words = candidate.split()
            if 1 <= len(cand_words) <= 5:
                filters["artist"] = candidate

    # Artist: "репертуар X" / "исполнитель X" / "группа X" / "артист X"
    # Handles Russian inflection: группа/группе/группы/группой/группу
    # Lookahead stops before question/continuation words so voice phrases like
    # "группу Электронный мальчик сколько песен у нас" extract just the name.
    if not filters.get("artist"):
        m = re.search(
            r"(?:репертуар|исполнител[ья]|групп[аеыуойи]?|артист[аеуыой]?|artist|band)"
            r"\s+([«»\w\s.,-]+?)"
            r"(?=\s*(?:[.,!?]|$"
            r"|\b(?:сколько|столько|как|какой|какая|какое|какие|каких"
            r"|где|когда|который|которые|которой|которого|которых"
            r"|у|есть|нет|имеется|имеются|числится|числятся"
            r"|там|здесь|тут|ведь|же|ли|то|что|чтобы|если|потому"
            r"|песен|треков|песни|трека|треке|альбомов|альбом"
            r"|я\s+посмотрю|посмотрю|посмотрим)\b))",
            text, re.IGNORECASE,
        )
        if m:
            artist = _strip_quotes(m.group(1).strip()).strip(".,!?;:")
            # Strip trailing noise: pronouns, perception verbs, prepositions, Excel requests
            artist = re.sub(
                r'\s+(?:ты|вы|он|она|они|я|мы|видишь|видите|вижу|видит|видно'
                r'|и|дай|скинь|в|на|из|как|за|с|по|до|от|excel|файл|xlsx'
                r'|сколько|столько|у|есть|нет|имеется|имеются|числится'
                r'|там|здесь|тут|песен|треков|песни|трека|треке'
                r').*$',
                '', artist, flags=re.IGNORECASE
            ).strip()
            # Discard if starts with a preposition — year-range context debris
            # e.g. "репертуар в периоде 1980-1989 годов" → after year strip → "в периоде  годов"
            if artist and re.match(r'^(?:в|на|из|за|с|по|до|от|у|при|под|над|для|со)\s', artist, re.IGNORECASE):
                artist = ""
            if artist:
                filters["artist"] = artist

    # Genre: "жанр шансон" / "стиль рок" / "категория джаз" or bare genre keyword
    m = _GENRE_CONTEXT_RE.search(text)
    if m:
        genre_word = m.group(1).strip().lower().rstrip(".,!?;:")
        filters["genre"] = GENRE_ALIASES.get(genre_word, genre_word)
    elif not filters.get("genre"):
        for alias_key, alias_val in GENRE_ALIASES.items():
            if re.search(r'\b' + re.escape(alias_key) + r'\b', lower):
                filters["genre"] = alias_val
                break

    # Alias lookup: single known label keyword (e.g. "Мелодия", "Зион")
    if not filters:
        text_clean = lower.strip(".,!?«» ")
        for alias_key, alias_val in LABEL_ALIASES.items():
            if alias_key in text_clean:
                filters["label"] = alias_val
                break

    # Fallback: bare words → artist only when no entity filter found yet.
    # Skip when year range already found AND text is long — avoids extracting
    # conversational garbage like "Ковальский табличку" from "репертуар в периоде 1980-1989 годов...".
    _year_found = bool(filters.get("year_from"))
    _text_long = len(text.split()) > 5
    if not filters.get("artist") and not filters.get("label") and not filters.get("genre") and not (_year_found and _text_long):
        stop = {
            "выгрузи", "выгрузим", "выгрузите", "выгружаем", "выгружаете", "выгружать",
            "выгрузка", "выгрузку", "выгружай", "выгрузить",
            "тысяч", "тысячи", "тысяча",
            "экспорт", "экспортируй", "сделай", "дай", "покажи", "скинь",
            "треки", "трек", "треков", "трека", "треке",
            "музыку", "музыка", "репертуар", "исполнитель",
            "группа", "из", "базы", "за", "год", "года", "период", "перид", "перио", "все", "мне",
            "пожалуйста", "список", "по", "полный", "полное", "полностью",
            "файл", "да", "нет", "всё", "отлично", "хорошо", "ладно",
            "хочу", "нужно", "нужен", "нужны", "можешь", "можно",
            "дайте", "отчёт", "отчет", "и", "а", "но", "в", "на", "для",
            "excel", "xlsx", "полная", "весь", "всю", "фирме", "фирма",
            "компании", "компания", "лейбле", "лейбла",
            "давай", "давайте", "конечно", "окей", "ок", "угу", "ага",
            "посмотри", "проверь", "напомни", "есть", "ли", "у", "нас",
            "что", "где", "когда", "какой", "какие", "сколько",
            "тебя", "тебе", "тобой", "твои", "твой", "твоя", "твоей",
            "базе", "базу", "базы", "базой",
            "каталоге", "каталогу", "каталог", "каталога",
            "реестре", "реестра", "реестру",
            "инфу", "информацию", "информации", "информация",
            "данные", "данных", "данным", "данного",
            "подбери", "подобрать", "подберёт",
            "песни", "песня", "песню",
            # Date/time words — not artist names
            "вчера", "сегодня", "позавчера", "завтра",
            "добавленные", "добавленных", "добавленным", "добавили", "добавлен",
            "загруженные", "загруженных", "загрузили", "загружен",
            "вышедшие", "вышедших", "вышли", "вышел", "вышедшим",
            "релиз", "релизе", "релиза", "release",
            # Year-period context words — not artist names
            "периоде", "периода", "периоду", "периодом", "периодах",
            "годов", "годах", "годами", "лет", "летних", "летний",
            # Nationality/descriptor adjectives — not artist names
            "русские", "русская", "русский", "русского", "русских", "русским", "русскую",
            "советские", "советская", "советский", "советского", "советских", "советским",
            "зарубежные", "зарубежная", "зарубежный", "зарубежных", "зарубежным",
            "отечественные", "отечественная", "отечественный",
            "популярные", "популярная", "популярный",
            "классические", "классическая", "классический",
            "старые", "старая", "старый", "старых",
            "новые", "новая", "новый", "новых",
            "знаешь", "знаете", "имеется", "имеются",
            "наш", "наша", "наше", "наши", "нашем", "нашей",
            "ты", "вы", "он", "она", "они", "я", "мы",
            "видишь", "видите", "вижу", "видит", "видно",
            "посмотрю", "посмотри", "посмотрим", "посмотрите", "посмотреть",
            "взгляну", "взгляни", "гляну", "загляну",
            "пришли", "пришлю", "сделай", "сделаю", "сделал",
            # demonstrative pronouns
            "этот", "эта", "это", "эти", "этой", "этого", "этому", "этим", "этих", "эту",
            "тот", "та", "то", "те", "той", "того", "тому", "тем", "тех", "ту",
            "данный", "данная", "данное", "данной", "данного",
            # noun case forms
            "группе", "группу", "группой",
            # pronoun oblique cases often leaked through ("что есть по ней")
            "ней", "нём", "ним", "ними", "них",
            "нам", "нами", "вам", "вами", "вас",
            "им", "ими", "их",
            "ему", "её", "его",
        }
        words = [_strip_quotes(w.strip(".,!?")) for w in text.split()
                 if w.lower().strip(".,!?«» ") not in stop
                 and not re.match(r'^(?:19|20)\d{2}(?:[-–—]\d{4})?$', w.strip(".,!?"))]
        words = [w for w in words if w]
        if 1 <= len(words) <= 4:
            filters["artist"] = " ".join(words)

    return filters


def fetch_tracks(filters: dict, limit: int = 5000) -> list[dict]:
    """Query Supabase tracks with the given filters. Returns list of track dicts."""
    params: dict = {
        "select": "id,title,artist,album,label,music_author,lyrics_author,release_date,genre_1,link",
        "order": "artist.asc,release_date.asc",
        "limit": str(limit),
    }

    conditions = []

    if filters.get("artist"):
        a = filters["artist"].replace("*", "").replace("(", "").replace(")", "")
        if " " in a:
            # Multi-word name: broad match is safe ("Группа Форум" won't clash)
            conditions.append(f"artist.ilike.*{a}*")
        else:
            # Single-word: word-boundary conditions so "Секрет" ≠ "Секретарь"
            # Values with spaces/commas MUST be double-quoted in PostgREST or()
            # syntax — unquoted commas are parsed as condition separators → 400.
            conditions.extend([
                f'artist.ilike."{a}"',       # exact
                f'artist.ilike."{a} *"',     # "Секрет feat. X"
                f'artist.ilike."* {a}"',     # "Группа Секрет"
                f'artist.ilike."* {a} *"',   # "Foo Секрет Bar"
                f'artist.ilike."{a},*"',     # "Секрет, Земляне"
                f'artist.ilike."* {a},*"',   # "Foo Секрет, Bar"
                f'artist.ilike."*,{a}"',     # "Земляне, Секрет"
                f'artist.ilike."*,{a} *"',   # "Foo, Секрет Bar"
            ])
        # Genitive/accusative ending normalization: "Киркорова" → try "Киркоров" too
        a_stripped = re.sub(r'[аяуюыиеёо]$', '', a, flags=re.IGNORECASE)
        if a_stripped != a and len(a_stripped) >= 4:
            if " " in a_stripped:
                conditions.append(f"artist.ilike.*{a_stripped}*")
            else:
                conditions.extend([
                    f'artist.ilike."{a_stripped}"',
                    f'artist.ilike."{a_stripped} *"',
                    f'artist.ilike."* {a_stripped}"',
                    f'artist.ilike."* {a_stripped} *"',
                ])

    if filters.get("label"):
        lb = filters["label"].replace("*", "")
        conditions.append(f"label.ilike.*{lb}*")
        # Normalize spaces around & so "S & P Digital" matches "S&P Digital"
        lb_compact = re.sub(r'\s*&\s*', '&', lb)
        if lb_compact != lb:
            conditions.append(f"label.ilike.*{lb_compact}*")

    if filters.get("genre"):
        g = filters["genre"].replace("*", "")
        conditions.append(f"genre_1.ilike.*{g}*")

    # Release-day filter: release_date contains DD.MM.YYYY string
    if filters.get("release_day"):
        try:
            rd = datetime.date.fromisoformat(filters["release_day"])
            conditions.append(f'release_date.ilike.*{rd.strftime("%d.%m.%Y")}*')
        except ValueError:
            pass

    # Year conditions — push into Supabase query directly so Latin-first
    # ORDER BY artist doesn't hide Cyrillic artists behind the row limit.
    # Without this, year-only queries post-filter on 5000 Latin-sorted rows
    # and never reach Cyrillic artists (А-Я, Unicode > Z).
    year_from = filters.get("year_from")
    year_to = filters.get("year_to", year_from)
    year_conds: list[str] = []
    if year_from is not None:
        for y in range(year_from, min(year_to + 1, year_from + 51)):
            year_conds.append(f"release_date.ilike.*{y}*")

    if conditions and year_conds:
        # Artist filter in DB via or=; year applied in Python post-filter below.
        # Avoids and=(or(),or()) nested syntax which Supabase PostgREST rejects.
        params["or"] = f"({','.join(conditions)})"
    elif conditions:
        params["or"] = f"({','.join(conditions)})"
    elif year_conds:
        # Year-only: push year conditions to DB so Cyrillic artists aren't
        # lost to Latin-first ORDER BY cutting off rows before limit.
        params["or"] = f"({','.join(year_conds)})"

    # Date-added filter: created_at range (requires duplicate param keys → list of tuples)
    date_params: list[tuple] = []
    if filters.get("date_added"):
        try:
            d = datetime.date.fromisoformat(filters["date_added"])
            next_d = d + datetime.timedelta(days=1)
            date_params = [("created_at", f"gte.{d.isoformat()}"), ("created_at", f"lt.{next_d.isoformat()}")]
        except ValueError:
            pass

    params_list = list(params.items()) + date_params
    r = httpx.get(f"{_sb_base()}/rest/v1/tracks", headers=_sb_headers(), params=params_list, timeout=45)
    r.raise_for_status()
    rows = r.json()

    # Always post-filter by year: fixes ilike false positives
    # (e.g. release_date="2004-2005" matches *2005* but first year = 2004)
    if year_from is not None:
        yt_safe = year_to or year_from
        rows = [
            row for row in rows
            if (m := re.search(r"\b((?:19|20)\d{2})\b", str(row.get("release_date") or "")))
            and year_from <= int(m.group(1)) <= yt_safe
        ]

    return rows


def fetch_recent_tracks(limit: int = 50) -> list[dict]:
    """Return the most recently added tracks ordered by id DESC."""
    params = {
        "select": "id,title,artist,album,label,music_author,lyrics_author,release_date,genre_1,link",
        "order": "id.desc",
        "limit": str(min(limit, 200)),
    }
    r = httpx.get(f"{_sb_base()}/rest/v1/tracks", headers=_sb_headers(), params=params, timeout=45)
    r.raise_for_status()
    return r.json()


def build_excel(tracks: list[dict], title: str = "Каталог SYNC LAB") -> bytes:
    """Build an .xlsx workbook from track list and return as bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Треки"

    # Header style
    header_fill = PatternFill(start_color="1F2D3D", end_color="1F2D3D", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    columns = [
        ("№", 5),
        ("Исполнитель", 30),
        ("Название", 40),
        ("Год", 8),
        ("Лейбл", 25),
        ("Альбом", 30),
        ("Автор музыки", 25),
        ("Автор текста", 25),
        ("Жанр", 15),
        ("Ссылка", 40),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = col_width

    ws.row_dimensions[1].height = 20

    def extract_year(s):
        if not s:
            return ""
        m = re.search(r"\b((?:19|20)\d{2})\b", str(s))
        return m.group(1) if m else str(s)

    for i, t in enumerate(tracks, 1):
        row = [
            i,
            t.get("artist") or "",
            t.get("title") or "",
            extract_year(t.get("release_date")),
            t.get("label") or "",
            t.get("album") or "",
            t.get("music_author") or "",
            t.get("lyrics_author") or "",
            t.get("genre_1") or "",
            t.get("link") or "",
        ]
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=i + 1, column=col_idx, value=value)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add a summary sheet
    ws_info = wb.create_sheet("Инфо")
    ws_info["A1"] = "Выгрузка"
    ws_info["B1"] = title
    ws_info["A2"] = "Треков"
    ws_info["B2"] = len(tracks)
    ws_info["A3"] = "Источник"
    ws_info["B3"] = "SYNC LAB / Supabase"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_export_caption(tracks: list[dict], subject: str) -> str:
    """Build a short caveman-style summary for the Excel file caption."""
    count = len(tracks)
    if not count:
        return f"{subject} — треков не найдено."

    albums = {}
    labels = set()
    years = set()
    for t in tracks:
        alb = t.get("album") or ""
        if alb:
            albums[alb] = albums.get(alb, 0) + 1
        lbl = t.get("label") or ""
        if lbl:
            labels.add(lbl)
        yr = re.search(r"\b((?:19|20)\d{2})\b", str(t.get("release_date") or ""))
        if yr:
            years.add(int(yr.group(1)))

    lines = [f"{subject} — {count} треков."]
    if albums:
        top = sorted(albums.items(), key=lambda x: -x[1])[:5]
        album_str = ", ".join(f"«{a}» ({n})" for a, n in top)
        if len(albums) > 5:
            album_str += f" + ещё {len(albums) - 5}"
        lines.append(f"Альбомы: {album_str}.")
    if labels:
        lines.append(f"Лейблы: {', '.join(sorted(labels)[:3])}.")
    if years:
        lines.append(f"Годы: {min(years)}–{max(years)}.")
    return "\n".join(lines)


def export_catalog(query_text: str) -> tuple[bytes, str, int, list[dict]]:
    """Parse query, fetch tracks, build Excel.

    Returns (xlsx_bytes, filename, track_count, tracks).
    """
    filters = parse_export_query(query_text)
    tracks = fetch_tracks(filters)

    # Bare-word fallback: artist search returned 0 → retry same term as label.
    # Handles "выгрузи Warner" / "выгрузи monolit" where no "лейбл" keyword given.
    if not tracks and filters.get("artist") and not filters.get("label") and not filters.get("genre"):
        label_filters = {k: v for k, v in filters.items() if k != "artist"}
        label_filters["label"] = filters["artist"]
        retry = fetch_tracks(label_filters)
        if retry:
            tracks = retry
            filters = label_filters

    # Build a human-readable filename
    parts = []
    if filters.get("artist"):
        safe = re.sub(r"[^a-zA-Zа-яёА-ЯЁ0-9 ]", "", filters["artist"])[:30].strip()
        parts.append(safe)
    if filters.get("label"):
        parts.append(filters["label"][:15])
    if filters.get("year_from"):
        yf = filters["year_from"]
        yt = filters.get("year_to", yf)
        parts.append(str(yf) if yf == yt else f"{yf}-{yt}")
    if filters.get("date_added"):
        parts.append(f"добавлено_{filters['date_added']}")
    if filters.get("release_day"):
        parts.append(f"релиз_{filters['release_day']}")
    suffix = "_".join(parts).replace(" ", "_") or "catalog"
    filename = f"SYNCLAB_{suffix}.xlsx"

    title_str = " | ".join(str(v) for v in filters.values() if v) or "Полный каталог"
    xlsx_bytes = build_excel(tracks, title=f"SYNC LAB — {title_str}")
    return xlsx_bytes, filename, len(tracks), tracks
